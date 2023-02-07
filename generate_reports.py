import os
import pandas as pd
import dataframe_image as dfi
import PySimpleGUI as sg
import datetime
import shutil
from openpyxl import load_workbook
import calendar

def generate_report(month):

    subsplashReport = pd.read_csv("Files/Drop Reports Here/"+month+" SS.csv")
    subsplashReport = subsplashReport[['first_name', 'last_name', 'subfund', 'net_amount', 'frequency', 'email']]
    subsplashReport.rename(columns={'first_name':'First Name', 'last_name':'Last Name', 'subfund':'Sub Fund', 'net_amount':'Net Amount', 'frequency':'Frequency', 'email':'Email'}, inplace=True)
    outputReport = subsplashReport

    try:
        planningCenterReport = pd.read_csv("Files/Drop Reports Here/"+month+" PCO.csv")
        planningCenterReport = planningCenterReport[['donor_first_name', 'donor_last_name', 'fund', 'net_amount']]
        planningCenterReport.rename(columns={'donor_first_name':"First Name", 'donor_last_name':"Last Name", 'fund':'Sub Fund', 'net_amount':'Net Amount'}, inplace=True)
        outputReport = pd.concat([planningCenterReport, subsplashReport])
    except:
        print("Planning Center Report Not Found")

    
    outputReport = outputReport.sort_values(by=['Sub Fund'])
    outputReport['Date'] = month

    for subFund in outputReport['Sub Fund'].unique():
        individualReport = outputReport.loc[outputReport['Sub Fund'] == subFund]
        #dfi.export(individualReport.style.hide(axis='index'), month + "/Individual Reports/" + subFund + " " + month + ".png")
        
        out_file = 'Files/Generated Reports/'+subFund+' Giving Report.xlsx'
        print(out_file)
        
        if(os.path.isfile(out_file)):
            with pd.ExcelWriter(out_file, mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer:
                try:
                    individualReport.to_excel(writer, sheet_name=month, startrow=writer.sheets[month].max_row, index=False)
                except:
                    individualReport.to_excel(writer, sheet_name=month, index=False)
                individualReport.to_excel(writer,sheet_name='Complete Giving List',index=False,header=False,startrow=writer.sheets[month].max_row)
            #open excel file
            #add to master giving log
            #add sheet to it
        else:
            shutil.copy("Dont Touch/template.xlsx", out_file)
            with pd.ExcelWriter(out_file, mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer:
                #reader = pd.read_excel(out_file, sheet_name='Complete Giving List')
                individualReport.to_excel(writer, sheet_name=month, index=False)
                individualReport.to_excel(writer,sheet_name= 'Complete Giving List',index=False,header=True)
            
            #create excel file and add the correct sheet
            
    outputReport.to_csv("Files/Generated Reports/" + month + " Staff Support Report.csv", index=False)

files = os.listdir(os.getcwd())
months = list(calendar.month_name)[1:]
month_year = [i+" "+str(datetime.date.today().year) for i in months]

layout = [[sg.Text("Please Select A Month:")], [sg.Combo(month_year)],[sg.Button("Submit")]]


window = sg.Window("CityLight Report Generator", layout, margins=(300,300))

while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED:
        break
    if event == "Submit":
        generate_report(values[0])
        break

window.close()

#write to excel spreadsheet for each person and have each sheet be a month for that person and the main sheet be an overview page
